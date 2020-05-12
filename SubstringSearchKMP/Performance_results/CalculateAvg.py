import openpyxl

path = 'Performance_condition2_DNA.xlsx'

wb = openpyxl.load_workbook(path)
listOfPages = wb.sheetnames
if not "Average" in listOfPages:
    wb.create_sheet("Average")
ws = wb["Average"]
ws1 = wb[listOfPages[0]]
ws2 = wb[listOfPages[1]]
ws3 = wb[listOfPages[2]]
temp = []


for cell in list(ws1.rows)[1]:
    temp.append(cell.value)

for i in range(len(temp)):
    ws.cell(2, i+1,temp[i])

temp.clear()

for i in [0, 1, 9, 10, 18, 19,27,28]:
    for cell in list(ws1.columns)[i]:
        temp.append(cell.value)
    for j in range(len(temp)):
        ws.cell(j + 1, i + 1, temp[j])
    temp.clear()

for i in (*range(2, 8),*range(11, 17),*range(20, 26),*range(29, 35)):
    for cell in list(ws1.columns)[i]:
        if (isinstance(cell.value, int)):
            wscell = ws.cell(cell.row, i + 1)
            # if isinstance(wscell.value, int):
            #     wscell.value = cell.value + wscell.value
            # else:
            wscell.value = (cell.value + ws2.cell(cell.row, i + 1).value + ws3.cell(cell.row, i + 1).value)/3



wb.save(path)
